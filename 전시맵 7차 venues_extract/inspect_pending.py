"""강남·서초 시트에서 검증완료 미입력 행을 운영여부별로 분류."""
from openpyxl import load_workbook

XLSX = "/Users/gosoogil/Downloads/서울미술공간_표본300_알바작업시트.xlsx"
wb = load_workbook(XLSX, read_only=True, data_only=True)
ws = wb["강남·서초"]

print("[검증완료 미입력 행 - 강남·서초]")
pending_rows = []
for r in range(5, 401):
    name = ws.cell(row=r, column=4).value
    if not name:
        continue
    operating = ws.cell(row=r, column=13).value or ""
    verified = ws.cell(row=r, column=17).value or ""
    if verified == "":  # 검증완료 미입력
        pending_rows.append({
            "row": r,
            "no": ws.cell(row=r, column=1).value,
            "name": name,
            "category": ws.cell(row=r, column=3).value,
            "address": ws.cell(row=r, column=5).value,
            "operating": operating,
            "homepage": ws.cell(row=r, column=11).value,
            "instagram": ws.cell(row=r, column=12).value,
        })

print(f"총 {len(pending_rows)}개\n")
by_op = {}
for row in pending_rows:
    by_op[row["operating"]] = by_op.get(row["operating"], 0) + 1
print("운영여부 분포:", by_op)
print()
for row in pending_rows:
    print(f"  No={row['no']}, 운영={row['operating']!r}, {row['category']}, {row['name']}")
    print(f"    주소: {row['address']}")
    print(f"    홈: {row['homepage']}  인스타: {row['instagram']}")
