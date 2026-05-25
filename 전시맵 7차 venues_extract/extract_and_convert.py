"""xlsx → venues.json 변환 스크립트

- 시트: 종로 / 중구 / 강남·서초
- 추출 조건: 운영여부=O 이고 (검증완료=OK 또는 미입력)
- 미입력은 verification_status="pending" 라벨로 구분
- 좌표는 Naver geocode가 채울 것이므로 비워둠
- 알려진 6곳 manual_coord_needed=True 표시
"""
import json
import re
import unicodedata
from openpyxl import load_workbook

XLSX = "/Users/gosoogil/Downloads/서울미술공간_표본300_알바작업시트.xlsx"
OUT_DIR = "/Users/gosoogil/Library/Mobile Documents/com~apple~CloudDocs/03 고수길 개인 자료/01 바이브코딩/전시맵 7차 venues_extract"

TARGET_SHEETS = ["종로", "중구", "강남·서초"]

# 컬럼 인덱스 (1-based)
COL = {
    "no": 1, "region_label": 2, "category": 3, "name": 4, "address": 5,
    "phone": 6, "email": 7, "source": 8, "score": 9, "auto_crawl": 10,
    "homepage": 11, "instagram": 12, "operating": 13, "coord_ok": 14,
    "hours": 15, "exhibit_channel": 16, "verified": 17, "memo": 18,
}
HEADER_ROW = 4
DATA_START = 5
DATA_END = 400  # 충분히 크게

# 권역 한국어 → 영문 subregion 매핑 (venue가 강남·서초 시트지만 실제 주소가 성동구인 경우 등 보정 필요)
SUBREGION_MAP = {
    "종로": "jongno",
    "중구": "junggu",
    "강남": "gangnam",
    "서초": "seocho",
    "성동": "seongdong",  # 일부 강남·서초 시트의 outlier
    "용산": "yongsan",
    "마포": "mapo",
}

CATEGORY_MAP = {
    "미술관": "미술관",
    "갤러리": "갤러리",
    "기타": "기타",
    "기타 문화공간": "기타",
}

# 핸드오프 문서에 언급된 Naver 오인식 6곳 (이름 일부 매칭으로 표시)
MANUAL_COORD_KEYWORDS = [
    "PS센터", "공간 형", "FF서울", "COSO", "갤러리모스", "서울미술관",
]


def slugify(name: str) -> str:
    """공간명 → venue_key (snake_case)."""
    if not name:
        return ""
    # 괄호 안 영어 별칭 추출 시도
    s = unicodedata.normalize("NFKC", name).strip()
    s = re.sub(r"[\(\[\{].*?[\)\]\}]", "", s)  # 괄호 제거
    s = s.strip()
    s = s.lower()
    # 영문/숫자만 남은 경우 그대로
    s = re.sub(r"[\s\-·\./]+", "_", s)
    s = re.sub(r"[^0-9a-zㄱ-힝_]", "", s)  # 한글 + a-z + 0-9 + _
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "unknown"


def english_alias(name: str) -> str | None:
    """공간명에서 괄호 안 영문 이름 추출."""
    if not name:
        return None
    m = re.search(r"[\(\[]([A-Za-z][A-Za-z0-9 \-_\.&']+)[\)\]]", name)
    return m.group(1).strip() if m else None


def base_name(name: str) -> str:
    """괄호·영문 별칭을 제거한 한국어 베이스 이름."""
    if not name:
        return ""
    s = re.sub(r"\s*[\(\[\{].*?[\)\]\}]\s*", "", name).strip()
    return s or name.strip()


def infer_subregion(sheet_name: str, address: str) -> str:
    """시트명 + 주소에서 subregion 추정."""
    addr = (address or "")
    # 주소 우선 매핑
    for k, v in SUBREGION_MAP.items():
        if f"{k}구" in addr:
            return v
    # 시트명 기반
    if sheet_name == "강남·서초":
        if "서초구" in addr:
            return "seocho"
        return "gangnam"
    return SUBREGION_MAP.get(sheet_name, sheet_name.lower())


def normalize_category(raw: str) -> str:
    if not raw:
        return "기타"
    raw = raw.strip()
    return CATEGORY_MAP.get(raw, "기타")


def needs_manual_coord(name: str) -> bool:
    if not name:
        return False
    return any(kw in name for kw in MANUAL_COORD_KEYWORDS)


def safe_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def clean_url(u: str) -> str | None:
    u = safe_str(u)
    if not u:
        return None
    if not u.lower().startswith(("http://", "https://")):
        return None
    return u


def extract_rows(ws, sheet_name: str):
    """시트에서 운영여부=O 이고 (검증완료 OK 또는 미입력)인 행 추출."""
    extracted = []
    skipped_stats = {"휴관": 0, "X": 0, "폐관": 0, "제외": 0, "진행중": 0, "운영여부빈값": 0}

    for r in range(DATA_START, DATA_END + 1):
        name = safe_str(ws.cell(row=r, column=COL["name"]).value)
        if not name:
            continue  # 빈 행

        operating = safe_str(ws.cell(row=r, column=COL["operating"]).value)
        verified = safe_str(ws.cell(row=r, column=COL["verified"]).value)

        # 운영여부 필터
        if operating == "":
            skipped_stats["운영여부빈값"] += 1
            continue
        if operating != "O":
            skipped_stats[operating] = skipped_stats.get(operating, 0) + 1
            continue

        # 검증완료 필터: OK 또는 미입력만 통과
        if verified == "제외":
            skipped_stats["제외"] += 1
            continue
        if verified == "진행중":
            skipped_stats["진행중"] += 1
            continue
        if verified not in ("OK", ""):
            # 알 수 없는 값은 일단 제외 + 카운트
            skipped_stats.setdefault(f"verified={verified}", 0)
            skipped_stats[f"verified={verified}"] += 1
            continue

        verification_status = "verified" if verified == "OK" else "pending"

        row = {
            "row_no": ws.cell(row=r, column=COL["no"]).value,
            "sheet": sheet_name,
            "region_label": safe_str(ws.cell(row=r, column=COL["region_label"]).value),
            "category_raw": safe_str(ws.cell(row=r, column=COL["category"]).value),
            "name": name,
            "address": safe_str(ws.cell(row=r, column=COL["address"]).value),
            "phone": safe_str(ws.cell(row=r, column=COL["phone"]).value),
            "email": safe_str(ws.cell(row=r, column=COL["email"]).value),
            "homepage": clean_url(ws.cell(row=r, column=COL["homepage"]).value),
            "instagram": clean_url(ws.cell(row=r, column=COL["instagram"]).value),
            "hours": safe_str(ws.cell(row=r, column=COL["hours"]).value),
            "exhibit_channel": safe_str(ws.cell(row=r, column=COL["exhibit_channel"]).value),
            "memo": safe_str(ws.cell(row=r, column=COL["memo"]).value),
            "verification_status": verification_status,
        }
        extracted.append(row)

    return extracted, skipped_stats


def to_venue(row: dict, seen_keys: set) -> dict:
    name = row["name"]
    base = base_name(name)
    alias_en = english_alias(name)
    aliases = []
    if alias_en:
        aliases.append(alias_en)
    if base != name:
        aliases.append(name)  # 괄호 포함 원본도 별칭으로

    # venue_key 고유성 보장
    key = slugify(alias_en) if alias_en else slugify(base)
    if not key or key == "unknown":
        key = slugify(name)
    orig = key
    suffix = 2
    while key in seen_keys:
        key = f"{orig}_{suffix}"
        suffix += 1
    seen_keys.add(key)

    venue = {
        "venue_key": key,
        "venue_name": base,
        "category": normalize_category(row["category_raw"]),
        "region": "seoul",
        "subregion": infer_subregion(row["sheet"], row["address"]),
        "address": row["address"],
        # lat/lng는 Naver geocode가 채움 — 일단 누락 표시
        "official_url": row["homepage"],
        "instagram_url": row["instagram"],
        "aliases": aliases,
        # 운영 정보 (extension fields)
        "phone": row["phone"] or None,
        "email": row["email"] or None,
        "hours": row["hours"] or None,
        "exhibit_channel": row["exhibit_channel"] or None,
        "verification_status": row["verification_status"],
    }

    if needs_manual_coord(name):
        venue["manual_coord_needed"] = True
        venue["_note"] = "Naver geocode 오인식 — 수동 좌표 입력 필요"

    # None 값 정리 (가독성)
    return {k: v for k, v in venue.items() if v not in (None, "", [])}


def main():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    all_venues = []
    seen_keys = set()
    per_sheet_stats = {}

    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"[!] '{sheet_name}' 시트 없음")
            continue
        ws = wb[sheet_name]
        rows, skipped = extract_rows(ws, sheet_name)

        sheet_venues = []
        for row in rows:
            sheet_venues.append(to_venue(row, seen_keys))

        per_sheet_stats[sheet_name] = {
            "extracted": len(sheet_venues),
            "verified_OK": sum(1 for v in sheet_venues if v.get("verification_status") == "verified"),
            "verification_pending": sum(1 for v in sheet_venues if v.get("verification_status") == "pending"),
            "skipped": skipped,
            "by_category": {},
        }
        for v in sheet_venues:
            c = v["category"]
            per_sheet_stats[sheet_name]["by_category"][c] = per_sheet_stats[sheet_name]["by_category"].get(c, 0) + 1
        all_venues.extend(sheet_venues)

    out_path = f"{OUT_DIR}/venues_new.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(all_venues, f, ensure_ascii=False, indent=2)

    report_path = f"{OUT_DIR}/extract_report.json"
    report = {
        "total_extracted": len(all_venues),
        "by_sheet": per_sheet_stats,
        "by_category_total": {},
        "by_verification_total": {},
        "by_subregion_total": {},
        "manual_coord_count": sum(1 for v in all_venues if v.get("manual_coord_needed")),
    }
    for v in all_venues:
        c = v["category"]
        report["by_category_total"][c] = report["by_category_total"].get(c, 0) + 1
        vs = v.get("verification_status", "?")
        report["by_verification_total"][vs] = report["by_verification_total"].get(vs, 0) + 1
        sr = v["subregion"]
        report["by_subregion_total"][sr] = report["by_subregion_total"].get(sr, 0) + 1
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    print(json.dumps(report, ensure_ascii=False, indent=2))
    print(f"\n→ venues_new.json: {out_path}")
    print(f"→ extract_report.json: {report_path}")


if __name__ == "__main__":
    main()
