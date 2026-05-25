"""xlsx 구조 파악: 시트 목록, 각 시트의 헤더 행과 컬럼 매핑, 데이터 시작 위치."""
import sys
from openpyxl import load_workbook

XLSX = "/Users/gosoogil/Downloads/서울미술공간_표본300_알바작업시트.xlsx"

EXPECTED_COLUMNS = [
    "No", "권역", "유형", "공간명", "주소", "전화", "이메일", "출처", "점수",
    "자동수집중", "홈페이지URL", "인스타URL", "운영여부", "좌표확인",
    "운영시간", "전시갱신채널", "검증완료", "비고",
]


def find_header_row(ws, max_search=10):
    """헤더 행 찾기 — '공간명' 또는 'No' 같은 키워드가 들어있는 첫 행."""
    for r in range(1, max_search + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 25)]
        joined = " ".join(str(v) for v in row_vals if v is not None)
        if "공간명" in joined or "venue" in joined.lower():
            return r, row_vals
    return None, None


def main():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    print("=" * 60)
    print(f"시트 목록 ({len(wb.sheetnames)}개):")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  - {name}  ({ws.max_row} rows × {ws.max_column} cols)")
    print("=" * 60)

    target_sheets = ["종로", "중구", "강남·서초"]
    for name in target_sheets:
        if name not in wb.sheetnames:
            # 점·중점·언더바 등 변형 케이스
            cand = [n for n in wb.sheetnames if name.replace("·", "") in n.replace("·", "")]
            print(f"\n[!] '{name}' 시트 직접 못 찾음. 후보: {cand}")
            if cand:
                name = cand[0]
            else:
                continue
        ws = wb[name]
        header_row, headers = find_header_row(ws)
        print(f"\n[{name}]")
        print(f"  헤더 행 = {header_row}")
        print(f"  헤더값 = {[h for h in headers if h is not None] if headers else None}")

        # 첫 데이터 행 미리보기 (헤더 + 1, + 2, + 3)
        if header_row:
            for off in range(1, 4):
                r = header_row + off
                vals = [ws.cell(row=r, column=c).value for c in range(1, 19)]
                if any(v not in (None, "") for v in vals):
                    print(f"  row {r}: {vals}")


if __name__ == "__main__":
    main()
